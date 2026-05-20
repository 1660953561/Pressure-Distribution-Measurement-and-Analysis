import pandas as pd
import numpy as np
import openpyxl
import os
from openpyxl.utils import get_column_letter

# 原始文件路径
input_file = r'D:\study\基于面压力分布在位测量的CMP加工机理研究\传感器数据\3kg5转\可使用分析数据\42\（10，11）\（10，11）.xlsx'

# 读取Excel文件
print("正在读取原始文件...")
df = pd.read_excel(input_file)

# 检查数据结构和列名
print(f"原始文件列名: {df.columns.tolist()}")
print(f"数据形状: {df.shape}")

# 获取第四列列名
pressure_column = df.columns[3]
print(f"将对第四列 '{pressure_column}' 进行转换...")

# 检查原始数据的统计信息
print("\n原始数据统计信息:")
print(f"最小值: {df[pressure_column].min()}")
print(f"最大值: {df[pressure_column].max()}")
print(f"平均值: {df[pressure_column].mean()}")
print(f"标准差: {df[pressure_column].std()}")

# 应用转换公式：P = 26273.4375 * exp(2.1978 * V)
def transform_pressure(V):
    """将Pressure值转换为新的P值"""
    #return 26273.4375 * np.exp(2.1978 * V)
    return 6.5662 * np.exp(2.1978 * V)

# 转换第四列
print("\n正在进行数据转换...")
df[pressure_column] = df[pressure_column].apply(transform_pressure)

# 显示转换后的统计信息
print("\n转换后的数据统计信息:")
print(f"最小值: {df[pressure_column].min()}")
print(f"最大值: {df[pressure_column].max()}")
print(f"平均值: {df[pressure_column].mean()}")
print(f"标准差: {df[pressure_column].std()}")

# 检查是否有异常值
print(f"\n转换后的数据类型: {df[pressure_column].dtype}")

# 创建新文件名
file_dir = os.path.dirname(input_file)
file_name = os.path.basename(input_file)
name_without_ext = os.path.splitext(file_name)[0]
output_file = os.path.join(file_dir, f"{name_without_ext}zhengti_N.xlsx")

# 保存为Excel文件（暂时不调整格式）
df.to_excel(output_file, index=False)

print(f"\n基本转换完成，文件已保存: {output_file}")

# 使用openpyxl调整Excel格式
print("正在优化Excel格式...")
wb = openpyxl.load_workbook(output_file)
ws = wb.active

# 1. 设置列宽（将第四列设置更宽）
ws.column_dimensions[get_column_letter(1)].width = 12  # Group_ID
ws.column_dimensions[get_column_letter(2)].width = 8   # Row
ws.column_dimensions[get_column_letter(3)].width = 8   # Col
ws.column_dimensions[get_column_letter(4)].width = 25  # Pressure - 设置更宽

# 2. 为第四列设置合适的数字格式
# 根据数值大小选择合适的格式
pressure_values = df[pressure_column]

# 检查数值范围
max_value = pressure_values.max()
print(f"\n转换后的最大压力值: {max_value}")

# 选择格式
if max_value > 1e6:  # 如果数值超过100万
    # 使用科学计数法
    number_format = '0.00E+00'
    print("使用科学计数法格式")
else:
    # 使用常规数字格式，保留4位小数
    number_format = '#,##0.0000'
    print("使用常规数字格式，保留4位小数")

# 应用格式到第四列
for row in range(2, ws.max_row + 1):
    cell = ws[f'D{row}']
    cell.number_format = number_format

# 3. 可选：添加标题样式
ws['A1'].font = openpyxl.styles.Font(bold=True)
ws['B1'].font = openpyxl.styles.Font(bold=True)
ws['C1'].font = openpyxl.styles.Font(bold=True)
ws['D1'].font = openpyxl.styles.Font(bold=True)

# 保存调整后的文件
wb.save(output_file)

print("\n" + "="*60)
print("转换完成!")
print(f"原始文件: {input_file}")
print(f"新文件: {output_file}")

# 读取转换后的文件进行验证
print("\n验证转换结果:")
df_check = pd.read_excel(output_file)
print(f"验证 - 文件形状: {df_check.shape}")
print(f"验证 - 前5行数据:")
print(df_check.head())

print(f"\n新文件路径: {os.path.abspath(output_file)}")
print(f"文件大小: {os.path.getsize(output_file)/1024:.2f} KB")
print("="*60)